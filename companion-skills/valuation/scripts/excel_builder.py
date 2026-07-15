"""
excel_builder.py — Institutional-style valuation workbook builder
=================================================================
Builds an Excel (.xlsx) workbook from a `results` dict produced by the
valuation toolkit (dcf.py, rnpv.py, relative_val.py, cost_of_capital.py).

Sheets created based on keys present in `results`:
  • Summary   — always
  • DCF       — method in {fcff, fcfe, ddm, apv} and results["dcf"] present
  • WACC      — method in {fcff, apv} and results["wacc_build"] present
  • rNPV_SOTP — method == "rnpv" and results["rnpv"] present
  • Comps     — method == "relative" and results["relative"] present

Requires: Python 3.8+, openpyxl
"""

from __future__ import annotations

import argparse
import math
import os
import sys
import tempfile
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side, numbers
    )
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
except ImportError as e:
    sys.exit(f"openpyxl is required: pip install openpyxl\n{e}")


# ---------------------------------------------------------------------------
# Colour constants (IB convention)
# ---------------------------------------------------------------------------
C_DARK_BLUE   = "1F4E79"   # section headers — white bold text
C_LIGHT_BLUE  = "D9E1F2"   # sub-headers
C_INPUT_FILL  = "F2F2F2"   # light grey — hardcoded inputs
C_OUTPUT_FILL = "BDD7EE"   # medium blue — key outputs / base case
C_WHITE       = "FFFFFF"
C_BLACK       = "000000"
C_INPUT_FONT  = "0000FF"   # blue font for input cells
C_FORMULA_FNT = "000000"   # black font for formula cells
C_LINK_FONT   = "008000"   # green font for cross-sheet links
C_OUTLIER_FNT = "FF0000"   # red for outlier peers

# Number formats
FMT_MILLIONS  = '#,##0.0'
FMT_THOUSANDS = '#,##0'
FMT_PCT       = '0.0%'
FMT_PER_SHARE = '#,##0.00'
FMT_MULTIPLE  = '0.0x'
FMT_TEXT      = '@'


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(color: str = C_DARK_BLUE) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold=False, color=C_BLACK, size=10, name="Calibri") -> Font:
    return Font(name=name, bold=bold, color=color, size=size)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _apply_section_header(ws, row: int, col: int, label: str,
                           merge_end_col: int = None, light=False):
    """Write a section header cell with dark/light blue fill and white/black bold text."""
    cell = ws.cell(row=row, column=col, value=label)
    if light:
        cell.fill    = _header_fill(C_LIGHT_BLUE)
        cell.font    = _font(bold=True, color=C_BLACK)
    else:
        cell.fill    = _header_fill(C_DARK_BLUE)
        cell.font    = _font(bold=True, color=C_WHITE)
    cell.alignment = _align(h="center")
    cell.border    = _thin_border()
    if merge_end_col and merge_end_col > col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=merge_end_col
        )
    return cell


def _input_cell(ws, row: int, col: int, value, fmt: str = FMT_MILLIONS):
    """Write a hardcoded input cell: light grey fill, blue font."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill           = _header_fill(C_INPUT_FILL)
    cell.font           = _font(color=C_INPUT_FONT)
    cell.number_format  = fmt
    cell.border         = _thin_border()
    cell.alignment      = _align(h="right")
    return cell


def _formula_cell(ws, row: int, col: int, formula: str, fmt: str = FMT_MILLIONS,
                  cross_sheet=False):
    """Write a formula cell: white fill, black (or green for cross-sheet) font."""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font           = _font(color=C_LINK_FONT if cross_sheet else C_FORMULA_FNT)
    cell.number_format  = fmt
    cell.border         = _thin_border()
    cell.alignment      = _align(h="right")
    return cell


def _output_cell(ws, row: int, col: int, value, fmt: str = FMT_MILLIONS):
    """Write a static key-output cell: medium blue fill, bold black font."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill           = _header_fill(C_OUTPUT_FILL)
    cell.font           = _font(bold=True)
    cell.number_format  = fmt
    cell.border         = _thin_border()
    cell.alignment      = _align(h="right")
    return cell


def _label_cell(ws, row: int, col: int, label: str, bold=False):
    cell = ws.cell(row=row, column=col, value=label)
    cell.font      = _font(bold=bold)
    cell.border    = _thin_border()
    cell.alignment = _align(h="left")
    return cell


def _set_col_widths(ws, widths: dict):
    """widths = {col_letter_or_int: width_in_chars}"""
    for col, w in widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w


def _freeze(ws, cell_ref: str):
    ws.freeze_panes = cell_ref


def _col(n: int) -> str:
    return get_column_letter(n)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, results: dict):
    ws = wb.create_sheet("Summary")
    company  = results.get("company", {})
    method   = results.get("method", "")
    warnings = results.get("warnings", [])
    pages    = results.get("methodology_pages", [])

    # Determine headline value
    headline = None
    headline_label = "Value per Share"
    dcf_data = results.get("dcf", {})
    rnpv_data = results.get("rnpv", {})
    rel_data  = results.get("relative", {})

    if method in ("fcff", "fcfe", "ddm", "apv") and dcf_data:
        headline = dcf_data.get("value_per_share")
    elif method == "rnpv" and rnpv_data:
        headline = rnpv_data.get("rounded_target") or rnpv_data.get("value_per_share")
        headline_label = "Scenario-Weighted Target"
    elif method == "relative" and rel_data:
        headline = rel_data.get("implied_per_share")
        headline_label = "Implied Per Share (Comps)"

    row = 1

    # ── Company header block ─────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "VALUATION SUMMARY", merge_end_col=6)
    row += 1

    info_pairs = [
        ("Company",  company.get("name", "")),
        ("Ticker",   company.get("ticker", "")),
        ("Method",   method.upper()),
        ("Currency", company.get("currency", "USD")),
        ("Units",    company.get("units", "millions")),
        ("As of",    company.get("as_of", "")),
    ]
    for label, val in info_pairs:
        _label_cell(ws, row, 1, label, bold=True)
        cell = ws.cell(row=row, column=2, value=val)
        cell.font   = _font()
        cell.border = _thin_border()
        row += 1

    row += 1

    # ── Headline value ───────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "HEADLINE VALUE", merge_end_col=6)
    row += 1
    _label_cell(ws, row, 1, headline_label, bold=True)
    cell = ws.cell(row=row, column=2, value=headline)
    cell.fill          = _header_fill(C_OUTPUT_FILL)
    cell.font          = _font(bold=True, size=14)
    cell.number_format = FMT_PER_SHARE
    cell.border        = _thin_border()
    cell.alignment     = _align(h="right")
    row += 2

    # ── Key Outputs table ────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "KEY OUTPUTS", merge_end_col=6)
    row += 1

    # Collect key outputs
    ko_rows: list[tuple] = []  # (label, value, fmt)

    if method in ("fcff", "fcfe", "ddm", "apv") and dcf_data:
        wacc_val = None
        if results.get("wacc_build"):
            wacc_val = results["wacc_build"].get("wacc")
        ko_rows += [
            ("Enterprise Value ($M)",   dcf_data.get("enterprise_value"), FMT_MILLIONS),
            ("Equity Value ($M)",       dcf_data.get("equity_value"),     FMT_MILLIONS),
            ("WACC",                    wacc_val or dcf_data.get("rate"), FMT_PCT),
            ("Terminal Growth Rate",    dcf_data.get("stable_growth"),    FMT_PCT),
            ("Terminal Value % of EV",  dcf_data.get("tv_pct"),           FMT_PCT),
            ("Value per Share",         dcf_data.get("value_per_share"),  FMT_PER_SHARE),
        ]
    elif method == "rnpv" and rnpv_data:
        ko_rows += [
            ("Pipeline Subtotal ($M)",  rnpv_data.get("pipeline_subtotal"), FMT_MILLIONS),
            ("Net Cash ($M)",           rnpv_data.get("net_cash"),          FMT_MILLIONS),
            ("Overhead PV ($M)",        rnpv_data.get("overhead_pv"),       FMT_MILLIONS),
            ("Equity Value ($M)",       rnpv_data.get("equity_value"),      FMT_MILLIONS),
            ("Shares (M)",              rnpv_data.get("shares"),            FMT_MILLIONS),
            ("Value per Share",         rnpv_data.get("value_per_share"),   FMT_PER_SHARE),
            ("Scenario-Weighted Target",rnpv_data.get("scenario_weighted"), FMT_PER_SHARE),
            ("Rounded Target",          rnpv_data.get("rounded_target"),    FMT_PER_SHARE),
        ]
    elif method == "relative" and rel_data:
        ko_rows += [
            ("Metric",                  rel_data.get("metric_name"),        FMT_TEXT),
            ("Target Metric ($M)",      rel_data.get("target_metric"),      FMT_MILLIONS),
            ("Median Multiple",         rel_data.get("median_multiple"),    FMT_MULTIPLE),
            ("Implied EV ($M)",         rel_data.get("implied_value"),      FMT_MILLIONS),
            ("Net Debt ($M)",           rel_data.get("net_debt"),           FMT_MILLIONS),
            ("Implied Per Share",       rel_data.get("implied_per_share"),  FMT_PER_SHARE),
        ]

    # Sub-header row
    _apply_section_header(ws, row, 1, "Metric",  light=True)
    _apply_section_header(ws, row, 2, "Value",   light=True)
    row += 1

    for label, val, fmt in ko_rows:
        _label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=val)
        c.font          = _font()
        c.number_format = fmt
        c.border        = _thin_border()
        c.alignment     = _align(h="right")
        row += 1

    row += 1

    # ── Methodology / Damodaran pages ─────────────────────────────────
    if pages:
        _apply_section_header(ws, row, 1, "METHODOLOGY / DAMODARAN REFERENCES",
                              merge_end_col=6)
        row += 1
        for pg in pages:
            cell = ws.cell(row=row, column=1, value=str(pg))
            cell.font      = _font()
            cell.border    = _thin_border()
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            row += 1
        row += 1

    # ── Warnings ─────────────────────────────────────────────────────────
    if warnings:
        _apply_section_header(ws, row, 1, "WARNINGS", merge_end_col=6)
        row += 1
        for w in warnings:
            cell = ws.cell(row=row, column=1, value=f"⚠ {w}")
            cell.font      = Font(name="Calibri", color="FF0000", size=10)
            cell.border    = _thin_border()
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            row += 1

    _set_col_widths(ws, {1: 36, 2: 22, 3: 18, 4: 18, 5: 18, 6: 18})
    _freeze(ws, "A2")
    return ws


def _build_dcf(wb: Workbook, results: dict):
    """Build the DCF sheet with live formulas."""
    dcf = results.get("dcf", {})
    if not dcf:
        return

    ws = wb.create_sheet("DCF")
    projection = dcf.get("projection", [])
    mid_year   = dcf.get("mid_year", True)
    driver_based = dcf.get("driver_based", False)  # FCFF built from revenue/margin/reinvestment
    company    = results.get("company", {})
    currency   = company.get("currency", "USD")

    row = 1

    # ── Title ────────────────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "DISCOUNTED CASH FLOW MODEL", merge_end_col=10)
    row += 2

    # ── Assumptions block (INPUT cells) ──────────────────────────────────
    _apply_section_header(ws, row, 1, "ASSUMPTIONS (INPUTS)", merge_end_col=4)
    row += 1

    assumptions: list[tuple] = [
        ("Base Cash Flow ($M)",   dcf.get("base_cf"),       FMT_MILLIONS),
        ("Discount Rate / WACC",  dcf.get("rate"),           FMT_PCT),
        ("Stable Growth Rate",    dcf.get("stable_growth"),  FMT_PCT),
        ("Risk-Free Rate",        dcf.get("rf"),             FMT_PCT),
        ("Net Debt ($M)",         dcf.get("net_debt", 0),    FMT_MILLIONS),
        ("Shares Outstanding (M)",dcf.get("shares"),         FMT_MILLIONS),
        ("Options Value ($M)",    dcf.get("options_value", 0), FMT_MILLIONS),
    ]

    # Track assumption cell addresses for formula references
    addr: dict[str, str] = {}
    asm_labels = [
        "base_cf", "rate", "stable_growth", "rf", "net_debt", "shares", "options_value"
    ]
    # Surface ROC when supplied so the value-creating-growth check's input is
    # auditable in the delivered model (display-only). [report D13]
    if dcf.get("roc") is not None:
        assumptions.append(("Return on Capital (ROC/ROE)", dcf.get("roc"), FMT_PCT))
        asm_labels.append("roc")

    for i, (label, val, fmt) in enumerate(assumptions):
        _label_cell(ws, row, 1, label, bold=False)
        ic = _input_cell(ws, row, 2, val, fmt)
        addr[asm_labels[i]] = f"$B${row}"
        row += 1

    row += 1
    base_cf_addr    = addr["base_cf"]
    rate_addr       = addr["rate"]
    sg_addr         = addr["stable_growth"]
    nd_addr         = addr["net_debt"]
    shares_addr     = addr["shares"]
    options_addr    = addr["options_value"]

    # ── Projection table ─────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "EXPLICIT PERIOD PROJECTION", merge_end_col=8)
    row += 1

    # Column headers
    proj_headers = ["Year", "Growth %", f"Cash Flow ({currency}M)",
                    "Period", "Discount Factor", f"PV ({currency}M)"]
    proj_cols = range(1, len(proj_headers) + 1)
    for c, h in zip(proj_cols, proj_headers):
        _apply_section_header(ws, row, c, h, light=True)
    row += 1

    proj_start_row = row  # first data row of projection

    # We write the growth and cash flow as literal values (inputs from projection dict)
    # then period, disc_factor, PV as formulas using the rate assumption cell
    cf_cells = []    # list of cell addresses for PV column (for SUM formula)

    for idx, p in enumerate(projection):
        yr  = p.get("year", idx + 1)
        g   = p.get("growth")
        cf  = p.get("cf")
        per = p.get("period")

        # Year — plain label
        ws.cell(row=row, column=1, value=yr).border = _thin_border()

        # Growth % — input (blue)
        if g is not None:
            _input_cell(ws, row, 2, g, FMT_PCT)
        else:
            ws.cell(row=row, column=2, value="").border = _thin_border()

        # Cash Flow. Driver-based FCFF (revenue×margin − reinvestment) is NOT a
        # geometric series, so write it as a literal input — PV/TV/EV formulas
        # still reference this cell, so the workbook matches results.json. Else use
        # the classic prior_cf*(1+growth) projection formula.
        if driver_based:
            cf_cell = _input_cell(ws, row, 3, cf, FMT_MILLIONS)
        elif idx == 0:
            cf_cell = _formula_cell(ws, row, 3, f"={base_cf_addr}*(1+{_col(2)}{row})", FMT_MILLIONS)
        else:
            cf_cell = _formula_cell(ws, row, 3, f"={_col(3)}{row - 1}*(1+{_col(2)}{row})", FMT_MILLIONS)
        cf_addr_here = f"{_col(3)}{row}"

        # Period — input (mid-year convention from projection dict or formula)
        if per is not None:
            _input_cell(ws, row, 4, per, '0.0')
        else:
            period_val = (idx + 0.5) if mid_year else (idx + 1)
            _input_cell(ws, row, 4, period_val, '0.0')

        # Discount factor — formula = 1 / (1 + rate)^period
        disc_formula = f"=1/(1+{rate_addr})^{_col(4)}{row}"
        _formula_cell(ws, row, 5, disc_formula, '0.000')

        # PV — formula = cash_flow * discount_factor
        pv_formula = f"={_col(3)}{row}*{_col(5)}{row}"
        pv_cell = _formula_cell(ws, row, 6, pv_formula, FMT_MILLIONS)
        cf_cells.append(f"{_col(6)}{row}")

        row += 1

    proj_end_row = row - 1

    row += 1

    # ── Terminal Value block ──────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "TERMINAL VALUE & BRIDGE", merge_end_col=8)
    row += 1

    # Last CF cell for terminal value calc
    last_cf_cell = f"{_col(3)}{proj_end_row}" if projection else base_cf_addr

    _label_cell(ws, row, 1, "Terminal Value ($M)")
    tv_cell_addr = f"{_col(2)}{row}"
    # Driver-based FCFF can use a stable reinvestment rate g/ROC (terminal_roc),
    # so the terminal cash flow is NOT last_cf*(1+g); the naive formula would
    # silently diverge from results.json. Write the authoritative Python terminal
    # value as a literal (the driver projection is literal-input anyway). Else the
    # classic Gordon formula. [verifier]
    if driver_based:
        _input_cell(ws, row, 2, dcf.get("terminal_value"), FMT_MILLIONS)
    else:
        _formula_cell(ws, row, 2, f"={last_cf_cell}*(1+{sg_addr})/({rate_addr}-{sg_addr})",
                      FMT_MILLIONS)
    row += 1

    # TV discount period = the LAST EXPLICIT period, which is exactly dcf.py's
    # authoritative exponent (n−0.5 under mid-year, n under end-year). The old
    # "+0.5 / +1" over-discounted the terminal value by half a period under the
    # default mid-year convention, so the workbook's PV(TV) ran ~4.4% below the
    # results.json value (~3.2% of total). Bare reference restores parity. [report D12]
    last_period_cell = f"{_col(4)}{proj_end_row}" if projection else None

    if last_period_cell:
        tv_period_formula = f"={last_period_cell}"
    else:
        tv_period_formula = "=1"

    _label_cell(ws, row, 1, "TV Discount Period")
    tv_period_cell_addr = f"{_col(2)}{row}"
    _formula_cell(ws, row, 2, tv_period_formula, '0.0')
    row += 1

    # PV(terminal) = TV / (1+rate)^tv_period
    pv_tv_formula = f"={tv_cell_addr}/(1+{rate_addr})^{tv_period_cell_addr}"
    _label_cell(ws, row, 1, "PV of Terminal Value ($M)")
    pv_tv_cell = f"{_col(2)}{row}"
    _formula_cell(ws, row, 2, pv_tv_formula, FMT_MILLIONS)
    row += 1

    # Σ explicit PV
    if cf_cells:
        sum_formula = f"=SUM({cf_cells[0]}:{cf_cells[-1]})"
    else:
        sum_formula = "=0"
    _label_cell(ws, row, 1, "Σ Explicit Period PV ($M)")
    sum_pv_cell = f"{_col(2)}{row}"
    _formula_cell(ws, row, 2, sum_formula, FMT_MILLIONS)
    row += 1

    # Enterprise value = sum_pv + pv_tv
    ev_formula = f"={sum_pv_cell}+{pv_tv_cell}"
    _label_cell(ws, row, 1, "Enterprise Value ($M)", bold=True)
    ev_cell = f"{_col(2)}{row}"
    c = _formula_cell(ws, row, 2, ev_formula, FMT_MILLIONS)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True)
    row += 1

    # TV % of EV
    tv_pct_formula = f"={pv_tv_cell}/{ev_cell}"
    _label_cell(ws, row, 1, "Terminal Value % of EV")
    _formula_cell(ws, row, 2, tv_pct_formula, FMT_PCT)
    row += 1

    # − Net Debt → Equity Value
    eq_formula = f"={ev_cell}-{nd_addr}-{options_addr}"
    _label_cell(ws, row, 1, "Equity Value ($M)", bold=True)
    eq_cell = f"{_col(2)}{row}"
    c = _formula_cell(ws, row, 2, eq_formula, FMT_MILLIONS)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True)
    row += 1

    # Value per share = equity / shares
    vps_formula = f"={eq_cell}/{shares_addr}*1"  # shares already in M
    _label_cell(ws, row, 1, "Value per Share", bold=True)
    vps_cell_dcf = f"{_col(2)}{row}"
    c = _formula_cell(ws, row, 2, vps_formula, FMT_PER_SHARE)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True, size=12)
    row += 2

    # ── Driver detail (revenue → margin → EBIT → reinvestment → FCFF) ──────
    # Makes the driver-built cash flows auditable alongside the FCFF inputs above.
    if driver_based and projection:
        _apply_section_header(ws, row, 1, "DRIVER DETAIL (Ch.22-23)", merge_end_col=8)
        row += 1
        dh = ["Year", "Revenue ($M)", "Op Margin", "EBIT ($M)",
              "EBIT(1−t) ($M)", "Reinvestment ($M)", "FCFF ($M)"]
        for ci, h in enumerate(dh, 1):
            _apply_section_header(ws, row, ci, h, light=True)
        row += 1
        for p in projection:
            ws.cell(row=row, column=1, value=p.get("year")).border = _thin_border()
            _input_cell(ws, row, 2, p.get("revenue"), FMT_MILLIONS)
            _input_cell(ws, row, 3, p.get("margin"), FMT_PCT)
            _input_cell(ws, row, 4, p.get("ebit"), FMT_MILLIONS)
            _input_cell(ws, row, 5, p.get("ebit_after_tax"), FMT_MILLIONS)
            _input_cell(ws, row, 6, p.get("reinvestment"), FMT_MILLIONS)
            _input_cell(ws, row, 7, p.get("cf"), FMT_MILLIONS)
            row += 1
        row += 1

    # ── Sensitivity Table ─────────────────────────────────────────────────
    sens = results.get("sensitivity")
    if sens:
        row_var  = sens.get("row_var", "WACC")
        col_var  = sens.get("col_var", "Terminal g")
        row_vals = sens.get("row_vals", [])
        col_vals = sens.get("col_vals", [])
        grid     = sens.get("grid", [])

        _apply_section_header(ws, row, 1,
            f"SENSITIVITY: {row_var} (rows) vs {col_var} (cols)",
            merge_end_col=1 + len(col_vals) + 1)
        row += 1

        # Header row — col_var values
        ws.cell(row=row, column=1, value=f"{row_var} \\ {col_var}").border = _thin_border()
        for ci, cv in enumerate(col_vals):
            c = ws.cell(row=row, column=2 + ci, value=cv)
            c.number_format = FMT_PCT
            c.fill   = _header_fill(C_LIGHT_BLUE)
            c.font   = _font(bold=True)
            c.border = _thin_border()
            c.alignment = _align(h="center")
        row += 1

        # Data rows — static values (no formula; grid is pre-computed)
        wacc_val = results.get("wacc_build", {}).get("wacc") if results.get("wacc_build") else dcf.get("rate")
        sg_val   = dcf.get("stable_growth")

        for ri, (rv, grid_row) in enumerate(zip(row_vals, grid)):
            rc = ws.cell(row=row, column=1, value=rv)
            rc.number_format = FMT_PCT
            rc.fill   = _header_fill(C_LIGHT_BLUE)
            rc.font   = _font(bold=True)
            rc.border = _thin_border()

            for ci, gv in enumerate(grid_row):
                is_base = (
                    (rv is not None and wacc_val is not None and
                     abs(rv - wacc_val) < 1e-9) and
                    (col_vals[ci] is not None and sg_val is not None and
                     abs(col_vals[ci] - sg_val) < 1e-9)
                )
                c = ws.cell(row=row, column=2 + ci, value=gv)
                c.number_format = FMT_PER_SHARE
                c.border = _thin_border()
                c.alignment = _align(h="right")
                if is_base:
                    c.fill = _header_fill(C_OUTPUT_FILL)
                    c.font = _font(bold=True)
            row += 1

    _set_col_widths(ws, {1: 32, 2: 18, 3: 18, 4: 12, 5: 18, 6: 18,
                         7: 18, 8: 18})
    _freeze(ws, "A2")
    return ws


def _build_wacc(wb: Workbook, results: dict):
    """Build the WACC sheet with live formulas."""
    wb_data = results.get("wacc_build", {})
    if not wb_data:
        return

    ws = wb.create_sheet("WACC")
    row = 1

    _apply_section_header(ws, row, 1, "WACC BUILD", merge_end_col=4)
    row += 2

    # ── Cost of Equity ────────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "COST OF EQUITY (CAPM)", merge_end_col=4, light=True)
    row += 1

    rf_row   = row
    _label_cell(ws, row, 1, "Risk-Free Rate (rf)")
    _input_cell(ws, row, 2, wb_data.get("rf"), FMT_PCT)
    row += 1

    beta_row = row
    _label_cell(ws, row, 1, "Levered Beta")
    _input_cell(ws, row, 2, wb_data.get("beta"), '0.00')
    row += 1

    erp_row  = row
    _label_cell(ws, row, 1, "Equity Risk Premium (ERP)")
    _input_cell(ws, row, 2, wb_data.get("erp"), FMT_PCT)
    row += 1

    _label_cell(ws, row, 1, "Cost of Equity (ke = rf + β×ERP)", bold=True)
    ke_formula = f"=$B${rf_row}+$B${beta_row}*$B${erp_row}"
    ke_cell = f"$B${row}"
    c = _formula_cell(ws, row, 2, ke_formula, FMT_PCT)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True)
    row += 2

    # ── Cost of Debt ──────────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "COST OF DEBT", merge_end_col=4, light=True)
    row += 1

    kd_row  = row
    _label_cell(ws, row, 1, "Cost of Debt (after-tax)")
    _input_cell(ws, row, 2, wb_data.get("cost_of_debt_after_tax"), FMT_PCT)
    kd_cell = f"$B${row}"
    row += 1

    tax_row = row
    _label_cell(ws, row, 1, "Tax Rate")
    _input_cell(ws, row, 2, wb_data.get("tax_rate"), FMT_PCT)
    row += 2

    # ── Capital Structure ─────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "CAPITAL STRUCTURE", merge_end_col=4, light=True)
    row += 1

    e_row = row
    _label_cell(ws, row, 1, "Market Value of Equity ($M)")
    _input_cell(ws, row, 2, wb_data.get("e_val"), FMT_MILLIONS)
    e_cell = f"$B${row}"
    row += 1

    d_row = row
    _label_cell(ws, row, 1, "Market Value of Debt ($M)")
    _input_cell(ws, row, 2, wb_data.get("d_val"), FMT_MILLIONS)
    d_cell = f"$B${row}"
    row += 1

    total_row = row
    _label_cell(ws, row, 1, "Total Capital ($M)")
    total_formula = f"={e_cell}+{d_cell}"
    total_cell = f"$B${row}"
    _formula_cell(ws, row, 2, total_formula, FMT_MILLIONS)
    row += 1

    _label_cell(ws, row, 1, "Weight of Equity (We)")
    we_formula = f"={e_cell}/{total_cell}"
    we_cell = f"$B${row}"
    _formula_cell(ws, row, 2, we_formula, FMT_PCT)
    row += 1

    _label_cell(ws, row, 1, "Weight of Debt (Wd)")
    wd_formula = f"={d_cell}/{total_cell}"
    wd_cell = f"$B${row}"
    _formula_cell(ws, row, 2, wd_formula, FMT_PCT)
    row += 2

    # ── WACC ──────────────────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "WACC", merge_end_col=4, light=True)
    row += 1

    wacc_formula = f"={ke_cell}*{we_cell}+{kd_cell}*{wd_cell}"
    _label_cell(ws, row, 1, "WACC = ke×We + kd×Wd", bold=True)
    c = _formula_cell(ws, row, 2, wacc_formula, FMT_PCT)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True, size=12)
    row += 1

    _set_col_widths(ws, {1: 38, 2: 18, 3: 18, 4: 18})
    _freeze(ws, "A2")
    return ws


def _build_rnpv(wb: Workbook, results: dict):
    """Build the rNPV_SOTP sheet."""
    rnpv = results.get("rnpv", {})
    if not rnpv:
        return

    ws = wb.create_sheet("rNPV_SOTP")
    assets    = rnpv.get("assets", [])
    scenarios = rnpv.get("scenarios", [])
    shares    = rnpv.get("shares")

    row = 1

    _apply_section_header(ws, row, 1,
        "rNPV SUM-OF-THE-PARTS VALUATION", merge_end_col=8)
    row += 2

    # ── Per-Asset Table ───────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "PIPELINE ASSETS", merge_end_col=8, light=False)
    row += 1

    asset_headers = ["Asset", "Peak Sales ($M)", "LoA (%)",
                     "PV Commercial ($M)", "PV Dev Cost ($M)",
                     "rNPV ($M)", "Per Share"]
    for ci, h in enumerate(asset_headers, 1):
        _apply_section_header(ws, row, ci, h, light=True)
    row += 1

    asset_start_row = row
    rnpv_cells = []

    for asset in assets:
        _label_cell(ws, row, 1, asset.get("name", ""))
        _input_cell(ws, row, 2, asset.get("peak_sales"), FMT_MILLIONS)
        _input_cell(ws, row, 3, asset.get("loa"), FMT_PCT)
        _input_cell(ws, row, 4, asset.get("pv_commercial"), FMT_MILLIONS)
        _input_cell(ws, row, 5, asset.get("pv_dev_cost"), FMT_MILLIONS)

        # rNPV = LoA*PV_commercial − PV_dev_cost, floored at 0 for optional programs
        # (right to abandon) but NOT for contractually `committed` ones, which carry
        # a negative rNPV — mirrors valuation_engine/asset_rnpv so the workbook agrees
        # with results.json (and with this row's own Per Share cell). [report D15]
        loa_cell   = f"{_col(3)}{row}"
        pvc_cell   = f"{_col(4)}{row}"
        pdc_cell   = f"{_col(5)}{row}"
        _core = f"{loa_cell}*{pvc_cell}-{pdc_cell}"
        rnpv_formula = f"={_core}" if asset.get("committed", False) else f"=MAX({_core},0)"
        rnpv_cell_addr = f"{_col(6)}{row}"
        _formula_cell(ws, row, 6, rnpv_formula, FMT_MILLIONS)
        rnpv_cells.append(rnpv_cell_addr)

        # Per share (needs shares from the totals section — write as static for now,
        # formula will ref shares input below)
        per_share_val = asset.get("per_share")
        c = ws.cell(row=row, column=7, value=per_share_val)
        c.number_format = FMT_PER_SHARE
        c.border = _thin_border()
        c.alignment = _align(h="right")

        row += 1

    asset_end_row = row - 1

    row += 1

    # ── Totals block ──────────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "EQUITY BRIDGE", merge_end_col=8, light=True)
    row += 1

    # Shares input cell
    shares_input_row = row
    _label_cell(ws, row, 1, "Shares Outstanding (M)")
    _input_cell(ws, row, 2, shares, FMT_MILLIONS)
    shares_cell = f"$B${row}"
    row += 1

    net_cash_row = row
    _label_cell(ws, row, 1, "+ Net Cash ($M)")
    _input_cell(ws, row, 2, rnpv.get("net_cash"), FMT_MILLIONS)
    net_cash_cell = f"$B${row}"
    row += 1

    overhead_row = row
    _label_cell(ws, row, 1, "− Overhead PV ($M)")
    _input_cell(ws, row, 2, rnpv.get("overhead_pv"), FMT_MILLIONS)
    overhead_cell = f"$B${row}"
    row += 1

    # Options value subtracted as a liability (never padded into the share count);
    # mirrors sotp_equity so the workbook matches results.json. [report D10]
    options_row = row
    _label_cell(ws, row, 1, "− Options Value ($M)")
    _input_cell(ws, row, 2, rnpv.get("options_value", 0.0), FMT_MILLIONS)
    options_cell = f"$B${row}"
    row += 1

    # Pipeline subtotal = SUM of rNPV cells
    if rnpv_cells:
        sub_formula = f"=SUM({rnpv_cells[0]}:{rnpv_cells[-1]})"
    else:
        sub_formula = "=0"
    _label_cell(ws, row, 1, "Pipeline Subtotal ($M)", bold=True)
    sub_cell = f"$B${row}"
    c = _formula_cell(ws, row, 2, sub_formula, FMT_MILLIONS)
    c.fill = _header_fill(C_OUTPUT_FILL)
    row += 1

    # Equity value = pipeline_sub + net_cash − overhead − options
    ev_formula = f"={sub_cell}+{net_cash_cell}-{overhead_cell}-{options_cell}"
    _label_cell(ws, row, 1, "Equity Value ($M)", bold=True)
    ev_cell = f"$B${row}"
    c = _formula_cell(ws, row, 2, ev_formula, FMT_MILLIONS)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True)
    row += 1

    # Value per share
    vps_formula = f"={ev_cell}/{shares_cell}"
    _label_cell(ws, row, 1, "Value per Share", bold=True)
    vps_cell = f"$B${row}"
    c = _formula_cell(ws, row, 2, vps_formula, FMT_PER_SHARE)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True, size=12)
    row += 2

    # ── Scenario Table ────────────────────────────────────────────────────
    if scenarios:
        _apply_section_header(ws, row, 1, "SCENARIO ANALYSIS", merge_end_col=8)
        row += 1

        scen_headers = ["Scenario", "Probability", "Price Target", "Weighted Value"]
        for ci, h in enumerate(scen_headers, 1):
            _apply_section_header(ws, row, ci, h, light=True)
        row += 1

        scen_prob_cells  = []
        scen_target_cells = []

        for sc in scenarios:
            _label_cell(ws, row, 1, sc.get("name", ""))
            _input_cell(ws, row, 2, sc.get("prob"), FMT_PCT)
            _input_cell(ws, row, 3, sc.get("target"), FMT_PER_SHARE)
            # weighted = prob * target
            prob_cell   = f"{_col(2)}{row}"
            target_cell = f"{_col(3)}{row}"
            wt_formula  = f"={prob_cell}*{target_cell}"
            _formula_cell(ws, row, 4, wt_formula, FMT_PER_SHARE)
            scen_prob_cells.append(prob_cell)
            scen_target_cells.append(f"{_col(4)}{row}")
            row += 1

        # Scenario-weighted sum
        if scen_target_cells:
            sw_formula = f"=SUM({scen_target_cells[0]}:{scen_target_cells[-1]})"
        else:
            sw_formula = "=0"
        _label_cell(ws, row, 1, "Scenario-Weighted Target", bold=True)
        sw_cell = f"$B${row}"
        c = _formula_cell(ws, row, 4, sw_formula, FMT_PER_SHARE)
        c.fill = _header_fill(C_OUTPUT_FILL)
        c.font = _font(bold=True)
        row += 1

        # Rounded target (static)
        rounded = rnpv.get("rounded_target") or rnpv.get("scenario_weighted")
        _label_cell(ws, row, 1, "Rounded Price Target", bold=True)
        c = ws.cell(row=row, column=4, value=rounded)
        c.fill          = _header_fill(C_OUTPUT_FILL)
        c.font          = _font(bold=True, size=12)
        c.number_format = FMT_PER_SHARE
        c.border        = _thin_border()
        c.alignment     = _align(h="right")
        row += 1

    _set_col_widths(ws, {1: 30, 2: 18, 3: 14, 4: 20, 5: 18, 6: 18, 7: 16, 8: 16})
    _freeze(ws, "A2")
    return ws


def _build_comps(wb: Workbook, results: dict):
    """Build the Comps sheet for relative valuation."""
    rel = results.get("relative", {})
    if not rel:
        return

    ws = wb.create_sheet("Comps")
    peers    = rel.get("peers", [])
    outliers = set(rel.get("outliers", []))
    metric   = rel.get("metric_name", "EV/EBITDA")
    shares   = rel.get("shares")

    row = 1

    _apply_section_header(ws, row, 1, f"COMPARABLE COMPANY ANALYSIS — {metric}",
                          merge_end_col=5)
    row += 2

    # ── Peer Table ────────────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "PEER TRADING MULTIPLES", merge_end_col=5, light=True)
    row += 1

    peer_headers = ["Company", f"{metric}", "Outlier?"]
    for ci, h in enumerate(peer_headers, 1):
        _apply_section_header(ws, row, ci, h, light=True)
    row += 1

    peer_mult_cells  = []  # for median formula
    peer_start_row   = row
    incl_mult_cells  = []

    for idx, peer in enumerate(peers):
        name = peer.get("name", f"Peer {idx+1}")
        mult = peer.get("multiple")
        is_out = idx in outliers

        name_cell = _label_cell(ws, row, 1, name)
        if is_out:
            name_cell.font = Font(name="Calibri", color=C_OUTLIER_FNT, size=10)

        mult_cell_addr = f"{_col(2)}{row}"
        mc = ws.cell(row=row, column=2, value=mult)
        mc.number_format = FMT_MULTIPLE
        mc.border        = _thin_border()
        mc.alignment     = _align(h="right")
        if is_out:
            mc.font = Font(name="Calibri", color=C_OUTLIER_FNT, size=10)

        out_cell = ws.cell(row=row, column=3,
                           value="excl." if is_out else "")
        out_cell.border    = _thin_border()
        out_cell.alignment = _align(h="center")
        if is_out:
            out_cell.font = Font(name="Calibri", color=C_OUTLIER_FNT, size=10)

        peer_mult_cells.append(mult_cell_addr)
        if not is_out:
            incl_mult_cells.append(mult_cell_addr)

        row += 1

    peer_end_row = row - 1

    row += 1

    # ── Valuation Bridge ─────────────────────────────────────────────────
    _apply_section_header(ws, row, 1, "APPLIED MULTIPLE & VALUATION", merge_end_col=5, light=True)
    row += 1

    # Target metric input
    _label_cell(ws, row, 1, f"Target {metric.split('/')[1] if '/' in metric else 'Metric'} ($M)")
    _input_cell(ws, row, 2, rel.get("target_metric"), FMT_MILLIONS)
    target_cell = f"$B${row}"
    row += 1

    # Net debt input
    _label_cell(ws, row, 1, "Net Debt ($M)")
    _input_cell(ws, row, 2, rel.get("net_debt", 0), FMT_MILLIONS)
    nd_cell = f"$B${row}"
    row += 1

    # Shares input (optional)
    if shares is not None:
        _label_cell(ws, row, 1, "Shares Outstanding (M)")
        _input_cell(ws, row, 2, shares, FMT_MILLIONS)
        shares_cell = f"$B${row}"
        row += 1
    else:
        shares_cell = None

    # Median multiple — use MEDIAN formula over included peers
    if incl_mult_cells:
        median_formula = f"=MEDIAN({incl_mult_cells[0]},{','.join(incl_mult_cells)})"
    elif peer_mult_cells:
        median_formula = f"=MEDIAN({peer_mult_cells[0]}:{peer_mult_cells[-1]})"
    else:
        median_formula = f"={rel.get('median_multiple', 0)}"

    _label_cell(ws, row, 1, "Median Multiple (excl. outliers)", bold=True)
    median_cell = f"$B${row}"
    c = _formula_cell(ws, row, 2, median_formula, FMT_MULTIPLE)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True)
    row += 1

    # Implied EV = target_metric * median_multiple
    ev_formula = f"={target_cell}*{median_cell}"
    _label_cell(ws, row, 1, "Implied Enterprise Value ($M)", bold=True)
    ev_cell = f"$B${row}"
    c = _formula_cell(ws, row, 2, ev_formula, FMT_MILLIONS)
    c.fill = _header_fill(C_OUTPUT_FILL)
    row += 1

    # Implied equity = EV - net_debt
    eq_formula = f"={ev_cell}-{nd_cell}"
    _label_cell(ws, row, 1, "Implied Equity Value ($M)", bold=True)
    eq_cell = f"$B${row}"
    c = _formula_cell(ws, row, 2, eq_formula, FMT_MILLIONS)
    c.fill = _header_fill(C_OUTPUT_FILL)
    c.font = _font(bold=True)
    row += 1

    # Per share (if shares available)
    if shares_cell:
        ps_formula = f"={eq_cell}/{shares_cell}"
        _label_cell(ws, row, 1, "Implied Per Share", bold=True)
        c = _formula_cell(ws, row, 2, ps_formula, FMT_PER_SHARE)
        c.fill = _header_fill(C_OUTPUT_FILL)
        c.font = _font(bold=True, size=12)
        row += 1

    # Low/High band if present
    low_val  = rel.get("low")
    high_val = rel.get("high")
    if low_val is not None and high_val is not None:
        row += 1
        _apply_section_header(ws, row, 1, "RANGE", merge_end_col=5, light=True)
        row += 1
        _label_cell(ws, row, 1, "Low")
        c = ws.cell(row=row, column=2, value=low_val)
        c.number_format = FMT_PER_SHARE
        c.border        = _thin_border()
        c.alignment     = _align(h="right")
        row += 1
        _label_cell(ws, row, 1, "High")
        c = ws.cell(row=row, column=2, value=high_val)
        c.number_format = FMT_MILLIONS
        c.border        = _thin_border()
        c.alignment     = _align(h="right")
        row += 1

    _set_col_widths(ws, {1: 32, 2: 20, 3: 12, 4: 18, 5: 18})
    _freeze(ws, "A2")
    return ws


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_workbook(results: dict, out_path: str) -> None:
    """
    Build an institutional-style Excel valuation workbook from `results` dict.

    Parameters
    ----------
    results : dict
        Valuation results following the schema in the module docstring.
    out_path : str
        Absolute path for the output .xlsx file.
    """
    wb = Workbook()
    # Remove the default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    method = results.get("method", "")

    # Always build Summary first
    _build_summary(wb, results)

    # DCF sheet
    if method in ("fcff", "fcfe", "ddm", "apv") and results.get("dcf"):
        _build_dcf(wb, results)
        if results.get("wacc_build"):
            _build_wacc(wb, results)

    # rNPV SOTP sheet
    if method == "rnpv" and results.get("rnpv"):
        _build_rnpv(wb, results)

    # Comps sheet
    if method == "relative" and results.get("relative"):
        _build_comps(wb, results)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------

def _make_fcff_results() -> dict:
    """Sample FCFF valuation results with projection + sensitivity + WACC."""
    return {
        "company": {
            "name": "Acme Corp",
            "ticker": "ACME",
            "currency": "USD",
            "units": "millions",
            "as_of": "2026-06-25",
        },
        "method": "fcff",
        "wacc_build": {
            "rf": 0.045,
            "erp": 0.055,
            "beta": 1.10,
            "cost_of_equity": 0.1055,
            "cost_of_debt_after_tax": 0.035,
            "tax_rate": 0.21,
            "e_val": 1200.0,
            "d_val": 300.0,
            "wacc": 0.0934,
        },
        "dcf": {
            "base_cf": 80.0,
            "stable_growth": 0.025,
            "rate": 0.0934,
            "rf": 0.045,
            "mid_year": True,
            "projection": [
                {"year": 1, "growth": 0.12, "cf": 89.6,  "period": 0.5, "disc_factor": 0.957, "pv": 85.8},
                {"year": 2, "growth": 0.10, "cf": 98.6,  "period": 1.5, "disc_factor": 0.875, "pv": 86.2},
                {"year": 3, "growth": 0.08, "cf": 106.4, "period": 2.5, "disc_factor": 0.799, "pv": 85.0},
                {"year": 4, "growth": 0.06, "cf": 112.8, "period": 3.5, "disc_factor": 0.731, "pv": 82.4},
                {"year": 5, "growth": 0.04, "cf": 117.3, "period": 4.5, "disc_factor": 0.668, "pv": 78.4},
            ],
            "terminal_value": 1750.0,
            "pv_terminal": 1169.0,
            "explicit_pv": 417.8,
            "enterprise_value": 1586.8,
            "tv_pct": 0.737,
            "net_debt": 200.0,
            "options_value": 0.0,
            "equity_value": 1386.8,
            "shares": 120.0,
            "value_per_share": 11.56,
        },
        "sensitivity": {
            "row_var": "WACC",
            "col_var": "Terminal g",
            "row_vals": [0.08, 0.0867, 0.0934, 0.10, 0.11],
            "col_vals": [0.015, 0.020, 0.025, 0.030, 0.035],
            "grid": [
                [14.2, 15.1, 16.3, 17.8, 19.8],
                [12.8, 13.5, 14.5, 15.7, 17.3],
                [11.4, 12.0, 12.9, 13.9, 15.2],
                [10.1, 10.7, 11.4, 12.3, 13.4],
                [ 8.6,  9.1,  9.6, 10.3, 11.1],
            ],
        },
        "warnings": ["Beta estimated from daily returns; consider weekly re-estimate."],
        "methodology_pages": [
            "[Damodaran, Investment Valuation 3e, p.300] — FCFF definition",
            "[p.312] — Mid-year discount convention",
            "[p.316] — Gordon Growth terminal value",
        ],
    }


def _make_rnpv_results() -> dict:
    """Sample rNPV / SOTP results (UPB-style biotech)."""
    # Assets: rNPV values 215, 150, 15 → pipeline subtotal = 380(ish via formula)
    # net_cash=255, overhead=110 → equity = 380+255-110 = 525
    # shares = 54.45 → per share = 525/54.45 ≈ 9.65
    return {
        "company": {
            "name": "UPB Pharma",
            "ticker": "UPB",
            "currency": "USD",
            "units": "millions",
            "as_of": "2026-06-25",
        },
        "method": "rnpv",
        "rnpv": {
            "assets": [
                {
                    "name": "UPB-101 (Lead, Ph3)",
                    "peak_sales": 1200.0,
                    "loa": 0.65,
                    "pv_commercial": 380.0,
                    "pv_dev_cost": 32.0,
                    "rnpv": 215.0,
                    "per_share": 3.95,
                },
                {
                    "name": "UPB-202 (Ph2)",
                    "peak_sales": 600.0,
                    "loa": 0.40,
                    "pv_commercial": 425.0,
                    "pv_dev_cost": 20.0,
                    "rnpv": 150.0,
                    "per_share": 2.76,
                },
                {
                    "name": "UPB-303 (Ph1)",
                    "peak_sales": 200.0,
                    "loa": 0.15,
                    "pv_commercial": 220.0,
                    "pv_dev_cost": 18.0,
                    "rnpv": 15.0,
                    "per_share": 0.28,
                },
            ],
            "pipeline_subtotal": 380.0,
            "net_cash": 255.0,
            "overhead_pv": 110.0,
            "equity_value": 525.0,
            "shares": 54.45,
            "value_per_share": 9.65,
            "scenarios": [
                {"name": "Bull",  "prob": 0.25, "target": 18.00},
                {"name": "Base",  "prob": 0.50, "target":  9.65},
                {"name": "Bear",  "prob": 0.25, "target":  4.25},
            ],
            "scenario_weighted": 10.39,
            "rounded_target":    10.00,
        },
        "methodology_pages": [
            "[Damodaran, Investment Valuation 3e, p.671] — rNPV for pharma/biotech",
            "[p.680] — Probability-weighted scenario analysis",
        ],
        "warnings": [],
    }


def _make_relative_results() -> dict:
    """Sample EV/EBITDA comps with outlier (peer index 3, multiple=40)."""
    return {
        "company": {
            "name": "Target Co",
            "ticker": "TGT",
            "currency": "USD",
            "units": "millions",
            "as_of": "2026-06-25",
        },
        "method": "relative",
        "relative": {
            "metric_name": "EV/EBITDA",
            "target_metric": 150.0,
            "peers": [
                {"name": "Peer A", "multiple": 8.0},
                {"name": "Peer B", "multiple": 9.0},
                {"name": "Peer C", "multiple": 11.0},
                {"name": "Peer D (outlier)", "multiple": 40.0},
            ],
            "median_multiple": 10.0,
            "implied_value": 1500.0,
            "low": 1200.0,
            "high": 1650.0,
            "outliers": [3],
            "net_debt": 200.0,
            "shares": 65.0,
            "implied_per_share": 20.0,
        },
        "methodology_pages": [
            "[Damodaran, Investment Valuation 3e, p.449] — EV/EBITDA multiples",
        ],
        "warnings": ["Peer D excluded as a strategic acquisition premium outlier."],
    }


def _parity_check() -> list:
    """Python==Excel terminal-value parity across conventions. [report D12]

    dcf.py discounts the terminal value at n−0.5 (mid-year) / n (end-year); the
    workbook must use the SAME exponent (the last explicit period). This verifies
    (a) the numeric equality of the Excel-convention PV(TV) with dcf.py's
    authoritative results.json value across n / rate / convention, and (b) that
    the emitted "TV Discount Period" formula is a bare period reference (no
    +0.5 / +1 that would re-introduce the half-period over-discount).
    """
    _here = os.path.dirname(os.path.abspath(__file__))
    if _here not in sys.path:
        sys.path.insert(0, _here)
    import dcf as _dcf
    fails = []
    for mid in (True, False):
        for growth in ([0.10, 0.08, 0.06], [0.12] * 5, [0.05] * 7):
            for rate in (0.08, 0.10, 0.12):
                res = _dcf.dcf_value(100.0, growth, rate, 0.025, mid_year=mid)
                n = len(growth)
                last_period = n - (0.5 if mid else 0.0)
                excel_pv_tv = res['tv_undiscounted'] / (1.0 + rate) ** last_period
                if abs(excel_pv_tv - res['terminal_pv']) > 1e-6:
                    fails.append(f"PV(TV) mismatch mid={mid} n={n} r={rate}: "
                                 f"excel={excel_pv_tv:.4f} dcf={res['terminal_pv']:.4f}")
    # Emitted-formula check: build a workbook and read the TV Discount Period cell.
    tmp = tempfile.mkdtemp(prefix="parity_")
    xp = os.path.join(tmp, "parity.xlsx")
    build_workbook(_make_fcff_results(), xp)
    ws = load_workbook(xp, data_only=False)["DCF"]
    tv_formula = None
    for r in ws.iter_rows():
        if isinstance(r[0].value, str) and "TV Discount Period" in r[0].value:
            tv_formula = r[1].value
            break
    if tv_formula is None:
        fails.append("TV Discount Period row not found in DCF sheet")
    elif '+' in str(tv_formula):
        fails.append(f"TV Discount Period still over-discounts (has '+'): {tv_formula!r}")
    return fails


def _rnpv_dilution_excel_check() -> list:
    """The delivered rNPV workbook must honor options-as-liability (D10) and the
    committed flag (D15), so the .xlsx agrees with results.json — and with itself.
    Guards against fixing only the Python side of a Python↔Excel pair.
    """
    fails = []
    res = {
        "company": {"name": "DilTest", "ticker": "DIL"}, "method": "rnpv",
        "rnpv": {"assets": [
            {"name": "lead", "pv_commercial": 1000.0, "loa": 1.0, "pv_dev_cost": 0.0,
             "committed": False, "rnpv": 1000.0, "per_share": 8.0},
            {"name": "committed_uw", "pv_commercial": 100.0, "loa": 0.3, "pv_dev_cost": 80.0,
             "committed": True, "rnpv": -50.0, "per_share": -0.5}],
            "pipeline_subtotal": 950.0, "net_cash": 0.0, "overhead_pv": 0.0,
            "options_value": 200.0, "equity_value": 750.0, "shares": 100.0,
            "value_per_share": 7.5},
        "warnings": [], "methodology_pages": [],
    }
    tmp = tempfile.mkdtemp(prefix="rnpvdil_")
    xp = os.path.join(tmp, "dil.xlsx")
    build_workbook(res, xp)
    ws = load_workbook(xp, data_only=False)["rNPV_SOTP"]
    options_row, equity_formula, committed_f, optional_f = None, None, None, None
    for r in ws.iter_rows():
        lab = r[0].value
        if isinstance(lab, str):
            if "Options Value" in lab:
                options_row = r[1].row
            if lab.strip().startswith("Equity Value"):
                equity_formula = r[1].value
            if lab == "committed_uw":
                committed_f = r[5].value      # col 6 = rNPV formula
            if lab == "lead":
                optional_f = r[5].value
    ef = str(equity_formula or "").replace("$", "")
    if options_row is None:
        fails.append("rNPV sheet has no '− Options Value' row — options not subtracted [D10]")
    elif f"B{options_row}" not in ef:
        fails.append(f"rNPV equity formula does not subtract options cell: {equity_formula!r} [D10]")
    if committed_f and "MAX" in str(committed_f).upper():
        fails.append(f"committed asset still floored with MAX in Excel: {committed_f!r} [D15]")
    if optional_f and "MAX" not in str(optional_f).upper():
        fails.append(f"optional asset lost its MAX floor in Excel: {optional_f!r} [D15]")
    return fails


def _driver_fcff_excel_check() -> list:
    """A driver-based FCFF workbook must render the FCFF stream as LITERAL inputs
    (revenue×margin − reinvestment is not geometric), so PV/TV/EV formulas match
    results.json, and must carry a DRIVER DETAIL block. Guards the 3-deliverable
    sync for the driver path.
    """
    _here = os.path.dirname(os.path.abspath(__file__))
    if _here not in sys.path:
        sys.path.insert(0, _here)
    import drivers as _dr
    fails = []
    rate, mid, g = 0.10, True, 0.03
    tax, tmargin, troc = 0.25, 0.15, 0.12
    dd = _dr.fcff_from_drivers(243.0, [0.6, 0.5, 0.4, 0.3, 0.25, 0.2, 0.15, 0.1, 0.08, 0.05],
                               0.0823, tmargin, 2.20, tax)
    # ROC-based stable-phase terminal value (what run_fcff_drivers computes with
    # terminal_roc) — deliberately DIFFERENT from the naive last_cf*(1+g)/(rate-g),
    # so this test would fail if the workbook reverted to the naive Gordon formula.
    rev_term = dd["revenues"][-1] * (1 + g)
    terminal_cf = rev_term * tmargin * (1 - tax) * (1 - g / troc)
    terminal_value = terminal_cf / (rate - g)
    naive_tv = dd["fcff"][-1] * (1 + g) / (rate - g)
    n = len(dd["fcff"])
    pv_terminal = terminal_value / (1 + rate) ** (n - 0.5)
    proj = []
    for i, cf in enumerate(dd["fcff"]):
        period = (i + 1) - 0.5
        proj.append({"year": i + 1, "revenue": dd["revenues"][i], "margin": dd["margins"][i],
                     "ebit": dd["ebit"][i], "ebit_after_tax": dd["ebit_after_tax"][i],
                     "reinvestment": dd["reinvestment"][i], "growth": None, "cf": cf,
                     "period": period, "disc_factor": 1.0 / (1 + rate) ** period,
                     "pv": cf / (1 + rate) ** period})
    ev = sum(p["pv"] for p in proj) + pv_terminal
    res = {"company": {"name": "Drv", "ticker": "DRV"}, "method": "fcff",
           "wacc_build": {"wacc": rate},
           "dcf": {"base_cf": dd["fcff"][0], "stable_growth": g, "rate": rate, "mid_year": mid,
                   "driver_based": True, "roc": troc, "projection": proj,
                   "terminal_value": terminal_value, "pv_terminal": pv_terminal,
                   "explicit_pv": sum(p["pv"] for p in proj), "enterprise_value": ev,
                   "tv_pct": pv_terminal / ev, "net_debt": 0.0, "equity_value": ev,
                   "shares": 100.0, "value_per_share": ev / 100.0, "options_value": 0.0,
                   "drivers": {"base_revenue": 243.0}},
           "warnings": [], "methodology_pages": []}
    tmp = tempfile.mkdtemp(prefix="drvexcel_")
    xp = os.path.join(tmp, "d.xlsx")
    build_workbook(res, xp)
    ws = load_workbook(xp, data_only=False)["DCF"]
    proj_cf, had_formula, detail, tv_cell = [], False, False, None
    for row in ws.iter_rows():
        if any(isinstance(c.value, str) and "DRIVER DETAIL" in c.value for c in row):
            detail = True
        if isinstance(row[0].value, str) and row[0].value.strip().startswith("Terminal Value ("):
            tv_cell = row[1].value
        yr = row[0].value
        per = row[3].value if len(row) > 3 else None
        if isinstance(yr, int) and isinstance(per, (int, float)) and abs(per - int(per) - 0.5) < 1e-9:
            v = row[2].value
            if isinstance(v, str) and v.startswith("="):
                had_formula = True
            else:
                proj_cf.append(v)
    if had_formula:
        fails.append("driver FCFF written as a geometric formula, not a literal (Excel≠results.json)")
    if len(proj_cf) != len(dd["fcff"]) or not all(abs(a - b) < 1e-6 for a, b in zip(proj_cf, dd["fcff"])):
        fails.append(f"driver FCFF literals mismatch: {proj_cf} vs {dd['fcff']}")
    if not detail:
        fails.append("DRIVER DETAIL section missing from the workbook")
    # Terminal value must be the authoritative ROC-based literal, not the naive formula.
    if abs(terminal_value - naive_tv) < 1.0:
        fails.append("test setup weak: ROC-based TV ≈ naive TV (cannot detect divergence)")
    elif isinstance(tv_cell, str) and tv_cell.startswith("="):
        fails.append("driver TV is a last_cf*(1+g)/(rate-g) formula — ignores terminal_roc (Excel≠results.json)")
    elif tv_cell is None:
        fails.append("Terminal Value row not found in the driver workbook")
    elif abs(tv_cell - terminal_value) > 1e-4:
        fails.append(f"driver TV literal {tv_cell:.4f} != results terminal_value {terminal_value:.4f}")
    return fails


def run_selftest(args):
    import traceback

    tmp_dir = tempfile.mkdtemp(prefix="valuation_selftest_")
    print(f"\nSelf-test output directory: {tmp_dir}\n")

    tests = [
        ("FCFF",     _make_fcff_results(),     "fcff_valuation.xlsx"),
        ("rNPV",     _make_rnpv_results(),      "rnpv_valuation.xlsx"),
        ("Relative", _make_relative_results(),  "relative_valuation.xlsx"),
    ]

    all_pass = True

    for name, res, fname in tests:
        out_path = os.path.join(tmp_dir, fname)
        print(f"─── {name} ({'method=' + res['method']}) ───")
        try:
            build_workbook(res, out_path)
            print(f"  [BUILD] OK → {out_path}")
        except Exception as e:
            traceback.print_exc()
            print(f"  [BUILD] FAIL — {e}")
            all_pass = False
            continue

        # Re-open and assert
        try:
            wb2 = load_workbook(out_path, data_only=False)
            sheet_names = wb2.sheetnames

            failures = []

            if name == "FCFF":
                # 1) Expected sheets
                for s in ("Summary", "DCF", "WACC"):
                    if s not in sheet_names:
                        failures.append(f"Missing sheet: {s}")

                # 2) Summary headline cell is numeric
                ws_sum = wb2["Summary"]
                # Headline is in B8 (after 6 info rows + 1 header + 1 headline row)
                # We scan for the first numeric cell in column B that looks like a per-share value
                headline_val = None
                for r in ws_sum.iter_rows(min_col=2, max_col=2):
                    for c in r:
                        if isinstance(c.value, (int, float)) and c.value is not None:
                            headline_val = c.value
                            break
                    if headline_val is not None:
                        break
                if headline_val is None:
                    failures.append("Summary: no numeric headline value found in col B")

                # 3) DCF sheet has at least one formula
                ws_dcf = wb2["DCF"]
                has_formula = False
                for r in ws_dcf.iter_rows():
                    for c in r:
                        if isinstance(c.value, str) and c.value.startswith("="):
                            has_formula = True
                            break
                    if has_formula:
                        break
                if not has_formula:
                    failures.append("DCF sheet: no formula cell found")

            elif name == "rNPV":
                # 1) Expected sheets
                for s in ("Summary", "rNPV_SOTP"):
                    if s not in sheet_names:
                        failures.append(f"Missing sheet: {s}")

                # 2) Summary equity value = 525 (static value in Key Outputs table)
                ws_sum = wb2["Summary"]
                found_525 = False
                for r in ws_sum.iter_rows(min_col=2, max_col=2):
                    for c in r:
                        if isinstance(c.value, (int, float)):
                            if abs(c.value - 525.0) < 0.5:
                                found_525 = True
                                break
                    if found_525:
                        break
                if not found_525:
                    failures.append("Summary rNPV: equity value 525 not found in col B "
                                    f"(searched all B cells)")

                # 3) rNPV sheet has formulas
                ws_rnpv = wb2["rNPV_SOTP"]
                has_formula = False
                for r in ws_rnpv.iter_rows():
                    for c in r:
                        if isinstance(c.value, str) and c.value.startswith("="):
                            has_formula = True
                            break
                    if has_formula:
                        break
                if not has_formula:
                    failures.append("rNPV_SOTP sheet: no formula cell found")

            elif name == "Relative":
                # 1) Expected sheets
                for s in ("Summary", "Comps"):
                    if s not in sheet_names:
                        failures.append(f"Missing sheet: {s}")

                # 2) Summary: implied_per_share = 20.0 present
                ws_sum = wb2["Summary"]
                found_20 = False
                for r in ws_sum.iter_rows(min_col=2, max_col=2):
                    for c in r:
                        if isinstance(c.value, (int, float)):
                            if abs(c.value - 20.0) < 0.5:
                                found_20 = True
                                break
                    if found_20:
                        break
                if not found_20:
                    failures.append("Summary Relative: implied_per_share 20.0 not found")

            if failures:
                for f in failures:
                    print(f"  [ASSERT] FAIL — {f}")
                all_pass = False
            else:
                print(f"  [ASSERT] PASS — all checks passed")

        except Exception as e:
            traceback.print_exc()
            print(f"  [ASSERT] FAIL — exception: {e}")
            all_pass = False

        print()

    print("─── Python==Excel TV parity [D12] ───")
    parity_fails = _parity_check()
    if parity_fails:
        for f in parity_fails:
            print(f"  [PARITY] FAIL — {f}")
        all_pass = False
    else:
        print("  [PARITY] PASS — TV discounted at last explicit period "
              "(n−0.5 mid / n end); workbook matches results.json")
    print()

    print("─── rNPV Excel dilution parity [D10/D15] ───")
    dil_fails = _rnpv_dilution_excel_check()
    if dil_fails:
        for f in dil_fails:
            print(f"  [DILUTION] FAIL — {f}")
        all_pass = False
    else:
        print("  [DILUTION] PASS — .xlsx subtracts options as a liability and "
              "carries committed assets unfloored (matches results.json)")
    print()

    print("─── Driver FCFF Excel parity [Ch.22-23] ───")
    drv_fails = _driver_fcff_excel_check()
    if drv_fails:
        for f in drv_fails:
            print(f"  [DRIVER] FAIL — {f}")
        all_pass = False
    else:
        print("  [DRIVER] PASS — driver FCFF rendered as literal inputs (workbook "
              "matches results.json) with a DRIVER DETAIL block")
    print()

    print("=" * 50)
    if all_pass:
        print("ALL TESTS PASSED")
        return 0
    else:
        print("ONE OR MORE TESTS FAILED")
        return 1


def main():
    parser = argparse.ArgumentParser(description="Valuation workbook builder")
    subparsers = parser.add_subparsers(dest="command")

    st_parser = subparsers.add_parser("selftest", help="Run self-test suite")

    args = parser.parse_args()

    if args.command == "selftest":
        sys.exit(run_selftest(args))
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
