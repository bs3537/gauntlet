#!/usr/bin/env python3
"""
validate_valuation.py — method-aware validation for the `valuation` skill.

Supersedes Anthropic's dcf-model/validate_dcf.py:
  * validates the engine's RESULTS JSON (recalc-independent, authoritative) for
    Damodaran guardrails across DCF / rNPV / relative methods, AND
  * optionally scans the .xlsx for Excel formula-error strings.
Fixes the upstream bug where a 'Sensitivity' sheet was required even though the
skill embeds sensitivity inside the DCF sheet.

Damodaran guardrails enforced (page anchors in references/):
  * terminal g < discount rate            [growth_and_terminal.md, p.306]
  * terminal g <= risk-free rate          [p.307]
  * terminal value 40-80% of EV (warn)    [dcf_fcff_fcfe_ddm.md]
  * WACC 5-20% (warn outside)             [cost_of_capital.md]
  * tax rate 15-30% (warn outside)
  * rNPV: every LoA in [0,1]; shares > 0  [rnpv_and_real_options.md]
  * relative: implied multiple within peer [low, high]; report outliers

CLI:
  python3 validate_valuation.py validate --results results.json [--excel model.xlsx] [--out out.json]
  python3 validate_valuation.py selftest
Exit 0 only when status == PASS.
"""
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, timezone

EXCEL_ERRORS = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
EXPECTED_SHEETS = {
    'fcff': ['DCF', 'WACC'], 'fcfe': ['DCF'], 'ddm': ['DCF'], 'apv': ['DCF', 'WACC'],
    'rnpv': ['rNPV_SOTP'], 'relative': ['Comps'],
}


class ValuationValidator:
    def __init__(self):
        self.errors, self.warnings, self.info = [], [], []

    # ---------- results-JSON validation (authoritative, recalc-independent) ----------
    def validate_results(self, r: dict):
        method = r.get('method', 'unknown')
        self.info.append(f"method: {method}")
        finite = lambda x: isinstance(x, (int, float)) and x == x and abs(x) != float('inf')

        if method in ('fcff', 'fcfe', 'ddm', 'apv'):
            self._validate_dcf(r, finite)
        elif method == 'rnpv':
            self._validate_rnpv(r, finite)
        elif method == 'relative':
            self._validate_relative(r, finite)
        else:
            self.warnings.append(f"Unknown method '{method}' — only structural checks run.")

        # universal: a headline value must exist and be finite
        headline = self._headline_value(r)
        if headline is None:
            self.errors.append("No finite headline value (value_per_share / equity_value / implied_value) found in results.")
        elif headline <= 0:
            self.warnings.append(f"Headline value is non-positive ({headline}). Verify inputs.")
        else:
            self.info.append(f"headline value: {headline:.4f}")

    def _headline_value(self, r):
        for sec in ('dcf', 'rnpv', 'relative'):
            d = r.get(sec) or {}
            for k in ('value_per_share', 'implied_per_share', 'equity_value', 'implied_value', 'scenario_weighted'):
                v = d.get(k)
                if isinstance(v, (int, float)) and v == v:
                    return v
        return None

    def _validate_dcf(self, r, finite):
        # `or {}` (not a default arg): DDM / direct-rate results carry
        # wacc_build == None, and r.get('wacc_build', {}) would return that
        # None and crash on .get() below. [report D1]
        d = r.get('dcf') or {}
        w = r.get('wacc_build') or {}
        rate = d.get('rate', w.get('wacc'))
        g = d.get('stable_growth')
        rf = d.get('rf', w.get('rf'))
        tv_pct = d.get('tv_pct')

        if finite(rate) and finite(g):
            if g >= rate:
                self.errors.append(f"CRITICAL: stable growth ({g:.2%}) >= discount rate ({rate:.2%}) — infinite/invalid terminal value [p.306].")
            else:
                self.info.append(f"OK: stable growth ({g:.2%}) < discount rate ({rate:.2%}).")
        if finite(rf) and finite(g) and g > rf + 1e-9:
            self.errors.append(f"CRITICAL: stable growth ({g:.2%}) > risk-free rate ({rf:.2%}) — violates terminal-growth cap [p.307].")
        # Growth must be paid for: g = RR × ROC, and growth only CREATES value when
        # ROC > the discount rate. When ROC is supplied, check both. [report D13]
        roc = d.get('roc')
        if finite(roc) and finite(g) and roc > 0:
            rr = g / roc
            if rr > 1.0 + 1e-9:
                self.warnings.append(f"Implied stable reinvestment rate g/ROC = {rr:.0%} exceeds 100% — growth {g:.2%} is infeasible at ROC {roc:.2%} [growth_and_terminal.md §4].")
            elif rr < -1e-9:
                self.warnings.append(f"Implied stable reinvestment rate g/ROC = {rr:.0%} is negative — check the signs of g and ROC.")
            else:
                self.info.append(f"OK: implied stable reinvestment rate g/ROC = {rr:.0%}.")
        if finite(roc) and finite(rate):
            if roc <= rate + 1e-9:
                self.warnings.append(f"ROC ({roc:.2%}) <= discount rate ({rate:.2%}) — growth is value-neutral-to-destructive; positive stable growth adds no intrinsic value [growth_and_terminal.md §4].")
            else:
                self.info.append(f"OK: ROC ({roc:.2%}) > discount rate ({rate:.2%}) — growth creates value.")
        # WACC band scales with a country-risk premium so legitimate high-CRP
        # (emerging-market) discount rates don't warn spuriously. [report D14]
        crp = w.get('crp') or 0.0
        wacc_hi = 0.20 + crp
        if finite(rate) and (rate < 0.05 or rate > wacc_hi):
            band = f"5–{wacc_hi:.0%}" + (f" (20% + {crp:.0%} country-risk premium)" if crp else "")
            self.warnings.append(f"WACC/discount rate ({rate:.2%}) outside {band} range — verify [cost_of_capital.md].")
        if finite(tv_pct):
            if tv_pct > 0.80:
                self.warnings.append(f"Terminal value is {tv_pct:.1%} of EV (>80%) — over-reliant on terminal assumptions.")
            elif tv_pct < 0.40:
                self.warnings.append(f"Terminal value is {tv_pct:.1%} of EV (<40%) — terminal assumptions may be too conservative.")
            else:
                self.info.append(f"OK: terminal value {tv_pct:.1%} of EV (40-80%).")
        tax = (r.get('wacc_build') or {}).get('tax_rate', (r.get('inputs') or {}).get('tax_rate'))
        tb = w.get('tax_band') or [0.15, 0.30]   # parameterizable per jurisdiction [report D14]
        # jsonschema's minItems/maxItems is never enforced at runtime, so a malformed
        # override must not crash or invert the check: require exactly two numeric,
        # ascending bounds, else fall back to the default band. [verifier]
        if not (isinstance(tb, (list, tuple)) and len(tb) == 2
                and all(isinstance(x, (int, float)) for x in tb) and tb[0] < tb[1]):
            tb = [0.15, 0.30]
        if finite(tax) and (tax < tb[0] or tax > tb[1]):
            self.warnings.append(f"Tax rate ({tax:.1%}) outside {tb[0]:.0%}-{tb[1]:.0%} range — verify.")

    def _validate_rnpv(self, r, finite):
        d = r.get('rnpv', {})
        for a in d.get('assets', []):
            loa = a.get('loa')
            if not (isinstance(loa, (int, float)) and 0 <= loa <= 1):
                self.errors.append(f"rNPV asset '{a.get('name','?')}': LoA {loa} not in [0,1] [rnpv guardrail].")
        shares = d.get('shares')
        if shares is not None and not (finite(shares) and shares > 0):
            self.errors.append(f"rNPV: shares ({shares}) must be > 0 for per-share value.")
        scen = d.get('scenarios') or []
        if scen:
            ptot = sum(s.get('prob', 0) for s in scen)
            if abs(ptot - 1.0) > 1e-6:
                self.warnings.append(f"Scenario probabilities sum to {ptot} (expected 1.0).")
            else:
                self.info.append("OK: scenario probabilities sum to 1.0.")

    def _validate_relative(self, r, finite):
        d = r.get('relative', {})
        mult = d.get('median_multiple')
        lo, hi = d.get('low_multiple'), d.get('high_multiple')
        if finite(mult) and finite(lo) and finite(hi) and not (lo <= mult <= hi):
            self.warnings.append(f"Applied multiple ({mult}) outside peer interquartile [{lo}, {hi}].")
        outs = d.get('outliers') or []
        if outs:
            if d.get('outliers_excluded'):
                self.info.append(f"Peer outliers at indices {outs} excluded from the central multiple.")
            else:
                # Do NOT claim exclusion that did not happen. [report D8]
                self.warnings.append(
                    f"Peer outliers flagged at indices {outs} but STILL INCLUDED in the "
                    f"central multiple — set relative.exclude_outliers to drop them.")

    # ---------- optional .xlsx scan ----------
    def validate_excel(self, path: str, method: str = None):
        try:
            import openpyxl
        except ImportError:
            self.warnings.append("openpyxl not installed — skipped .xlsx scan.")
            return
        p = Path(path)
        if not p.exists():
            self.warnings.append(f".xlsx not found: {path} — skipped scan.")
            return
        wb_v = openpyxl.load_workbook(path, data_only=True)
        wb_f = openpyxl.load_workbook(path, data_only=False)
        # expected sheets (method-aware; NO phantom 'Sensitivity' requirement)
        if method and method in EXPECTED_SHEETS:
            for s in EXPECTED_SHEETS[method]:
                if s in wb_v.sheetnames:
                    self.info.append(f"Found sheet: {s}")
                else:
                    self.warnings.append(f"Recommended sheet missing for {method}: {s}")
        # formula-error scan + recalc-availability detection
        total_formulas, cached_present, found = 0, 0, 0
        for sn in wb_v.sheetnames:
            wv, wf = wb_v[sn], wb_f[sn]
            for row in wf.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        total_formulas += 1
                        cv = wv[cell.coordinate].value
                        if cv is not None:
                            cached_present += 1
                        if isinstance(cv, str):
                            for err in EXCEL_ERRORS:
                                if err in cv:
                                    self.errors.append(f"{err} at {sn}!{cell.coordinate}")
                                    found += 1
                                    break
        self.info.append(f"Formulas in workbook: {total_formulas}")
        if total_formulas and cached_present == 0:
            self.info.append("Recalc unavailable (no cached formula values) — formula-error scan limited; open in Excel to recalc. Logic validated from results JSON.")
        elif found == 0:
            self.info.append("No Excel formula errors found in cached values.")

    def result(self, files):
        return {
            'files': files,
            'validation_date': datetime.now(timezone.utc).isoformat(),
            'status': 'PASS' if not self.errors else 'FAIL',
            'error_count': len(self.errors),
            'warning_count': len(self.warnings),
            'errors': self.errors, 'warnings': self.warnings, 'info': self.info,
        }


def run_validate(results_path, excel_path=None, out_path=None):
    v = ValuationValidator()
    method = None
    r = json.loads(Path(results_path).read_text())
    method = r.get('method')
    v.validate_results(r)
    if excel_path:
        v.validate_excel(excel_path, method)
    res = v.result({'results': results_path, 'excel': excel_path})
    print(json.dumps(res, indent=2))
    if out_path:
        Path(out_path).write_text(json.dumps(res, indent=2))
    return res


def selftest():
    checks = []
    # 1. clean FCFF results -> PASS
    good = {'method': 'fcff', 'wacc_build': {'wacc': 0.09, 'rf': 0.04, 'tax_rate': 0.24},
            'dcf': {'rate': 0.09, 'stable_growth': 0.025, 'rf': 0.04, 'tv_pct': 0.62,
                    'equity_value': 1200.0, 'value_per_share': 24.0}}
    v = ValuationValidator(); v.validate_results(good)
    checks.append(("clean FCFF -> no errors", len(v.errors) == 0))
    # 2. terminal g >= rate -> FAIL
    bad = json.loads(json.dumps(good)); bad['dcf']['stable_growth'] = 0.10
    v = ValuationValidator(); v.validate_results(bad)
    checks.append(("g>=rate -> error", any('>= discount rate' in e for e in v.errors)))
    # 3. terminal g > rf -> FAIL
    bad2 = json.loads(json.dumps(good)); bad2['dcf']['stable_growth'] = 0.05  # > rf 0.04, < rate 0.09
    v = ValuationValidator(); v.validate_results(bad2)
    checks.append(("g>rf -> error", any('risk-free' in e for e in v.errors)))
    # 4. rNPV bad LoA -> FAIL ; good UPB -> PASS
    upb = {'method': 'rnpv', 'rnpv': {'assets': [{'name': 'asthma', 'loa': 0.50}, {'name': 'crswnp', 'loa': 0.58}],
            'shares': 54.45, 'scenarios': [{'name': 'bull', 'prob': 0.25, 'target': 18},
            {'name': 'base', 'prob': 0.45, 'target': 9.65}, {'name': 'bear', 'prob': 0.30, 'target': 4.25}],
            'value_per_share': 9.65, 'scenario_weighted': 10.1175}}
    v = ValuationValidator(); v.validate_results(upb)
    checks.append(("UPB rNPV -> no errors", len(v.errors) == 0))
    badloa = json.loads(json.dumps(upb)); badloa['rnpv']['assets'][0]['loa'] = 1.4
    v = ValuationValidator(); v.validate_results(badloa)
    checks.append(("LoA>1 -> error", any('not in [0,1]' in e for e in v.errors)))
    # 5. relative outlier surfaces
    rel = {'method': 'relative', 'relative': {'median_multiple': 10.0, 'low_multiple': 8.5, 'high_multiple': 11.0,
            'outliers': [3], 'implied_value': 5000.0, 'implied_per_share': 50.0}}
    v = ValuationValidator(); v.validate_results(rel)
    checks.append(("relative clean -> no errors", len(v.errors) == 0))

    # 6. malformed tax_band overrides must never crash or invert the check — all
    #    fall back to the default [0.15,0.30] band, under which a 40% rate warns. [verifier/D14]
    tb_ok = True
    for bad in ([0.20], [], [0.30, 0.15], ["a", "b"], [None, None], "x", {'lo': 0.1}):
        v = ValuationValidator()
        try:
            v.validate_results({'method': 'fcff',
                                'wacc_build': {'wacc': 0.09, 'tax_rate': 0.40, 'tax_band': bad},
                                'dcf': {'rate': 0.09, 'stable_growth': 0.025, 'value_per_share': 24.0}})
            if not any('Tax rate' in w for w in v.warnings):
                tb_ok = False
        except Exception:
            tb_ok = False
    checks.append(("malformed tax_band never crashes; falls back to default band", tb_ok))

    # reversed band must fall back, NOT spuriously warn on a normal 20% tax
    v = ValuationValidator()
    v.validate_results({'method': 'fcff',
                        'wacc_build': {'wacc': 0.09, 'tax_rate': 0.20, 'tax_band': [0.30, 0.15]},
                        'dcf': {'rate': 0.09, 'stable_growth': 0.025, 'value_per_share': 24.0}})
    checks.append(("reversed tax_band falls back (no spurious warn on 20% tax)",
                   not any('Tax rate' in w for w in v.warnings)))

    # 7. country-risk premium widens the WACC band (no spurious EM warning) [D14]
    v = ValuationValidator()
    v.validate_results({'method': 'fcff', 'wacc_build': {'wacc': 0.24, 'crp': 0.07},
                        'dcf': {'rate': 0.24, 'stable_growth': 0.03, 'value_per_share': 10.0}})
    checks.append(("CRP widens WACC band (no warn at 24% with 7% CRP)",
                   not any('WACC/discount rate' in w for w in v.warnings)))

    print("=" * 60); print("validate_valuation.py selftest")
    ok = True
    for name, passed in checks:
        print(f"  [{'PASS' if passed else 'FAIL'}] {name}")
        ok = ok and passed
    print("=" * 60); print("ALL PASS" if ok else "SOME FAILED")
    return ok


def main():
    ap = argparse.ArgumentParser(description="Method-aware valuation validator")
    sub = ap.add_subparsers(dest='cmd')
    pv = sub.add_parser('validate'); pv.add_argument('--results', required=True)
    pv.add_argument('--excel'); pv.add_argument('--out')
    sub.add_parser('selftest')
    args = ap.parse_args()
    if args.cmd == 'validate':
        res = run_validate(args.results, args.excel, args.out)
        sys.exit(0 if res['status'] == 'PASS' else 1)
    elif args.cmd == 'selftest':
        sys.exit(0 if selftest() else 1)
    else:
        ap.print_help(); sys.exit(1)


if __name__ == '__main__':
    main()
